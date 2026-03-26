"""Utilities for loading tabular files and running groupml experiments."""

from __future__ import annotations

import json
from importlib.util import find_spec
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd

from .config import GroupMLConfig
from .result import GroupMLResult
from .runner import GroupMLRunner, compare_group_strategies
from .summaries import build_summary_payload, build_summary_tables, summary_text


_CSV_FALLBACK_ENCODINGS: tuple[str, ...] = ("utf-8", "utf-8-sig", "cp1252", "latin-1")


def _excel_number_format_for_column(column_name: str) -> str | None:
    lowered = str(column_name).strip().lower()
    if any(
        token in lowered
        for token in ["cv_", "test_", "score", "rmse", "mae", "metric", "error", "improvement", "std"]
    ):
        return "0.00000"
    if lowered.startswith("n_") or lowered.endswith("_count") or lowered in {"runs", "rank", "fold", "row_index"}:
        return "0"
    return None


def _format_openpyxl_worksheet(writer: Any, sheet_name: str, df: pd.DataFrame) -> None:
    if not excel_export_available():
        return
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None:
        return
    if worksheet.max_row >= 2:
        worksheet.freeze_panes = "A2"

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, column_name in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        series = df[column_name] if column_name in df.columns else pd.Series(dtype=object)
        if not series.empty:
            preview = series.tolist()
            max_len = max([len(str(column_name))] + [len(str(value)) for value in preview[:200]])
        else:
            max_len = len(str(column_name))
        worksheet.column_dimensions[letter].width = max(10, min(60, max_len + 2))

        number_format = _excel_number_format_for_column(str(column_name))
        if number_format is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=idx)
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = number_format


def load_tabular_data(path: str | Path, **read_kwargs: Any) -> pd.DataFrame:
    """Load a CSV or Excel file into a pandas DataFrame."""
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        if "encoding" in read_kwargs and read_kwargs["encoding"] is not None:
            return pd.read_csv(file_path, **read_kwargs)

        last_error: UnicodeDecodeError | None = None
        for encoding in _CSV_FALLBACK_ENCODINGS:
            try:
                return pd.read_csv(file_path, encoding=encoding, **read_kwargs)
            except UnicodeDecodeError as exc:
                last_error = exc
                continue
        if last_error is not None:
            raise last_error
        return pd.read_csv(file_path, **read_kwargs)
    if suffix in {".xls", ".xlsx"}:
        return pd.read_excel(file_path, **read_kwargs)
    raise ValueError(
        f"Unsupported file extension '{file_path.suffix}'. "
        "Supported: .csv, .xls, .xlsx"
    )


def fit_evaluate_file(
    path: str | Path,
    config: GroupMLConfig,
    callbacks: Iterable[Callable[[dict[str, Any]], None]] | None = None,
    **read_kwargs: Any,
) -> GroupMLResult:
    """Load a file and run `GroupMLRunner.fit_evaluate`."""
    df = load_tabular_data(path, **read_kwargs)
    return GroupMLRunner(config).fit_evaluate(df, callbacks=callbacks)


def compare_group_strategies_file(
    path: str | Path,
    target: str,
    feature_columns: list[str] | None = None,
    group_columns: list[str] | None = None,
    rule_splits: list[str] | None = None,
    read_kwargs: dict[str, Any] | None = None,
    callbacks: Iterable[Callable[[dict[str, Any]], None]] | None = None,
    **kwargs: Any,
) -> GroupMLResult:
    """Functional API wrapper for running experiments from a tabular file."""
    df = load_tabular_data(path, **(read_kwargs or {}))
    return compare_group_strategies(
        df=df,
        target=target,
        feature_columns=feature_columns,
        group_columns=group_columns,
        rule_splits=rule_splits,
        callbacks=callbacks,
        **kwargs,
    )


def default_report_filename(prefix: str = "groupml_report", ext: str = ".csv") -> str:
    """Build a friendly timestamped report filename."""
    clean_ext = ext if ext.startswith(".") else f".{ext}"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}{clean_ext}"


def default_summary_filename(prefix: str = "groupml_summary", ext: str = ".csv") -> str:
    """Build a friendly timestamped summary filename."""
    return default_report_filename(prefix=prefix, ext=ext)


def excel_export_available() -> bool:
    """Return True when openpyxl is available for xlsx export."""
    return find_spec("openpyxl") is not None


def preferred_tabular_extension(report_format: str = "auto") -> str:
    """Resolve preferred extension for tabular exports."""
    fmt = str(report_format).strip().lower()
    if fmt == "csv":
        return ".csv"
    if fmt == "excel":
        return ".xlsx"
    return ".xlsx" if excel_export_available() else ".csv"


def export_report(
    result: GroupMLResult,
    path: str | Path,
    sheet_name: str = "leaderboard",
) -> Path:
    """Export leaderboard report to CSV or Excel based on file extension."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()

    if suffix == ".csv":
        result.leaderboard.to_csv(output_path, index=False)
        return output_path

    if suffix in {".xls", ".xlsx"}:
        try:
            with pd.ExcelWriter(output_path) as writer:
                result.leaderboard.to_excel(writer, index=False, sheet_name=sheet_name)
                _format_openpyxl_worksheet(writer, sheet_name, result.leaderboard)
        except ImportError as exc:
            raise ImportError(
                "Excel export requires an Excel writer engine (for example openpyxl). "
                "Install it or export as .csv."
            ) from exc
        return output_path

    raise ValueError(
        f"Unsupported report extension '{output_path.suffix}'. Supported: .csv, .xls, .xlsx"
    )


def export_raw_report(
    result: GroupMLResult,
    path: str | Path,
    sheet_name: str = "raw_report",
) -> Path:
    """Export raw per-row report to CSV or Excel based on file extension."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()
    raw_df = result.raw_report if isinstance(result.raw_report, pd.DataFrame) else pd.DataFrame()

    if suffix == ".csv":
        raw_df.to_csv(output_path, index=False)
        return output_path

    if suffix in {".xls", ".xlsx"}:
        try:
            with pd.ExcelWriter(output_path) as writer:
                raw_df.to_excel(writer, index=False, sheet_name=sheet_name)
                _format_openpyxl_worksheet(writer, sheet_name, raw_df)
        except ImportError as exc:
            raise ImportError(
                "Excel export requires an Excel writer engine (for example openpyxl). "
                "Install it or export as .csv."
            ) from exc
        return output_path

    raise ValueError(
        f"Unsupported raw report extension '{output_path.suffix}'. Supported: .csv, .xls, .xlsx"
    )


def export_summary(
    result: GroupMLResult,
    path: str | Path,
    top_n: int = 10,
    sheet_name: str = "summary",
) -> Path:
    """Export human-friendly summary to text/JSON/CSV/Excel."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()

    if suffix in {".txt", ".md"}:
        output_path.write_text(summary_text(result, top_n=top_n), encoding="utf-8")
        return output_path

    if suffix == ".json":
        payload = build_summary_payload(result, top_n=top_n)
        output_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        return output_path

    if suffix == ".csv":
        tables = build_summary_tables(result, top_n=top_n)
        stacked_tables: list[pd.DataFrame] = []
        for section, table in tables.items():
            if table.empty:
                continue
            with_section = table.copy()
            if "section" in with_section.columns:
                with_section.insert(0, "table", section)
            else:
                with_section.insert(0, "section", section)
            stacked_tables.append(with_section)
        if stacked_tables:
            summary_csv = pd.concat(stacked_tables, ignore_index=True, sort=False)
        else:
            summary_csv = pd.DataFrame({"section": []})
        summary_csv.to_csv(output_path, index=False)
        return output_path

    if suffix in {".xls", ".xlsx"}:
        tables = build_summary_tables(result, top_n=top_n)
        try:
            with pd.ExcelWriter(output_path) as writer:
                ordered_items = [(name, table) for name, table in tables.items() if name != "warnings"]
                if "warnings" in tables:
                    ordered_items.append(("warnings", tables["warnings"]))
                for idx, (name, table) in enumerate(ordered_items):
                    if idx == 0:
                        safe_name = sheet_name[:31]
                    else:
                        safe_name = name[:31]
                    table.to_excel(writer, index=False, sheet_name=safe_name)
                    _format_openpyxl_worksheet(writer, safe_name, table)
        except ImportError as exc:
            raise ImportError(
                "Excel export requires an Excel writer engine (for example openpyxl). "
                "Install it or export summary as .txt/.json/.csv."
            ) from exc
        return output_path

    raise ValueError(
        f"Unsupported summary extension '{output_path.suffix}'. Supported: .txt, .md, .json, .csv, .xls, .xlsx"
    )


def export_reporting_bundle(
    result: GroupMLResult,
    path: str | Path,
    top_n: int = 10,
    report_format: str = "auto",
    include_raw: bool = True,
) -> dict[str, Path]:
    """Export summary/warnings/all runs/raw results as one workbook or split CSV files."""
    base_path = Path(path)
    base_path.parent.mkdir(parents=True, exist_ok=True)
    summary_tables = build_summary_tables(result, top_n=top_n)
    summary_df = summary_tables.get("summary", pd.DataFrame())
    recommendations_df = summary_tables.get("recommendations", pd.DataFrame())
    warnings_df = summary_tables.get("warnings", pd.DataFrame(columns=["warning"]))
    runs_df = result.all_runs.copy() if isinstance(result.all_runs, pd.DataFrame) and not result.all_runs.empty else result.leaderboard.copy()
    raw_df = result.raw_report.copy() if include_raw and isinstance(result.raw_report, pd.DataFrame) else pd.DataFrame()

    fmt = str(report_format).strip().lower()
    if fmt not in {"auto", "excel", "csv"}:
        raise ValueError("report_format must be one of: auto, excel, csv.")
    if fmt == "auto":
        suffix = base_path.suffix.lower()
        if suffix in {".xlsx", ".xls"}:
            fmt = "excel"
        elif suffix == ".csv":
            fmt = "csv"
        else:
            fmt = "excel" if excel_export_available() else "csv"

    if fmt == "excel" and not excel_export_available():
        fmt = "csv"

    if fmt == "excel":
        output_path = base_path if base_path.suffix.lower() in {".xlsx", ".xls"} else base_path.with_suffix(".xlsx")
        with pd.ExcelWriter(output_path) as writer:
            summary_df.to_excel(writer, index=False, sheet_name="summary")
            _format_openpyxl_worksheet(writer, "summary", summary_df)
            recommendations_df.to_excel(writer, index=False, sheet_name="recommendations")
            _format_openpyxl_worksheet(writer, "recommendations", recommendations_df)
            runs_df.to_excel(writer, index=False, sheet_name="all_runs")
            _format_openpyxl_worksheet(writer, "all_runs", runs_df)
            if include_raw:
                raw_df.to_excel(writer, index=False, sheet_name="raw_results")
                _format_openpyxl_worksheet(writer, "raw_results", raw_df)
            warnings_df.to_excel(writer, index=False, sheet_name="warnings")
            _format_openpyxl_worksheet(writer, "warnings", warnings_df)
        outputs: dict[str, Path] = {
            "workbook": output_path,
            "summary": output_path,
            "recommendations": output_path,
            "warnings": output_path,
            "all_runs": output_path,
        }
        if include_raw:
            outputs["raw_results"] = output_path
        return outputs

    summary_path = base_path if base_path.suffix.lower() == ".csv" else base_path.with_suffix(".csv")
    warnings_path = summary_path.with_name(f"{summary_path.stem}_warnings.csv")
    recommendations_path = summary_path.with_name(f"{summary_path.stem}_recommendations.csv")
    runs_path = summary_path.with_name(f"{summary_path.stem}_all_runs.csv")
    raw_path = summary_path.with_name(f"{summary_path.stem}_raw_results.csv")

    summary_df.to_csv(summary_path, index=False)
    warnings_df.to_csv(warnings_path, index=False)
    recommendations_df.to_csv(recommendations_path, index=False)
    runs_df.to_csv(runs_path, index=False)
    outputs = {
        "summary": summary_path,
        "recommendations": recommendations_path,
        "warnings": warnings_path,
        "all_runs": runs_path,
    }
    if include_raw:
        raw_df.to_csv(raw_path, index=False)
        outputs["raw_results"] = raw_path
    return outputs
